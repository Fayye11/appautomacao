import PySimpleGUI as sg
import openpyxl

sg.theme('Reddit')

# Layout da plataforma
layout = [
    [sg.Text('Saldo inicial', size=(12, 0)), sg.Input(size=(20, 0), key='2')],
    [sg.Text('Saldo Final', size=(12, 0)), sg.Input(size=(20, 0), key='3')],
    [sg.Button('Salvar', key='Salvar')]
]

window = sg.Window('Cadastro de Produtos', layout)


workbook = openpyxl.load_workbook('PLANILHATRADERESSENTIALS.xlsx')
planilha = workbook['SET']

def encontrar_proxima_linha_vazia():
    for row in range(3, planilha.max_row + 2):  
        if planilha.cell(row=row, column=2).value is None: 
            return row
    return planilha.max_row + 1

# Loop principal
while True:
    event, values = window.read()

    if event == sg.WIN_CLOSED:  
        break

    elif event == 'Salvar':  
        saldo_inicial = values['2']
        saldo_final = values['3']
        
        if not saldo_inicial or not saldo_final:
            sg.popup('Por favor, preencha todos os campos')
        else:
            
            next_row = encontrar_proxima_linha_vazia()

            planilha.cell(row=next_row, column=4).value = saldo_inicial  
            planilha.cell(row=next_row, column=5).value = saldo_final 
           
            workbook.save('PLANILHATRADERESSENTIALS.xlsx')
           
            sg.popup('Alteração feita com sucesso!')

            window['2'].update('')
            window['3'].update('')

window.close()
